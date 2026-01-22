import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
import pyperclip
from openpyxl import load_workbook, Workbook
import tempfile, os
import tkinter.font as tkFont
import math
from tkinter import PhotoImage, Label

# Try to import gseapy for ORA; show a friendly message if missing
try:
    import gseapy as gp
    _GSEAPY_AVAILABLE = True
except Exception:
    _GSEAPY_AVAILABLE = False

# Appologies for lack of comments

class GeneComparerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🧬ATPlease Work!!!🧬")
        self.root.geometry("900x600")

        # ===== Global Font =====
        default_font = tkFont.nametofont("TkDefaultFont")
        default_font.configure(family="Segoe UI", size=10)

        # ===== Dark Theme Styling =====
        self.root.configure(bg="#0b1d3a")  # deep navy
        style = ttk.Style()
        style.theme_use("clam")

        # Notebook tabs
        style.configure("TNotebook", background="#0b1d3a", borderwidth=0)
        style.configure("TNotebook.Tab", background="#1a2f5a", foreground="white",
                        font=("Segoe UI", 10, "bold"), padding=[10, 5])
        style.map("TNotebook.Tab", background=[("selected", "#2e4a87")])

        # Buttons
        style.configure("TButton", background="#1a2f5a", foreground="white",
                        font=("Segoe UI", 10, "bold"), padding=6, borderwidth=0)
        style.map("TButton", background=[("active", "#2e4a87")])

        # Combobox
        style.configure("TCombobox", fieldbackground="#1a2f5a", background="#1a2f5a",
                        foreground="white", arrowcolor="white", selectbackground="#2e4a87",
                        font=("Segoe UI", 10))

        # Labels
        style.configure("TLabel", background="#0b1d3a", foreground="white", font=("Segoe UI", 10))

        # Treeview
        style.configure("Treeview", background="#1a2f5a", foreground="white",
                        fieldbackground="#1a2f5a", font=("Segoe UI", 10))
        style.map("Treeview", background=[("selected", "#2e4a87")], foreground=[("selected", "white")])
        style.configure("Treeview.Heading", background="#0b1d3a", foreground="white",
                        font=("Segoe UI", 10, "bold"))

        # ===== Spinner ===== (not working)
        self.progress = ttk.Progressbar(self.root, mode="indeterminate")

        # ===== Data Vars =====
        self.gene_file = None
        self.data_file = None
        self.gene_df = None
        self.data_df = None

        # ===== Frames for Page Switching =====
        self.main_frame = tk.Frame(self.root, bg="#0b1d3a")
        self.pathway_frame = tk.Frame(self.root, bg="#0b1d3a")
        self.main_frame.pack(fill="both", expand=True)  # show main page first

        # ===== Menu Bar (View -> switch pages) =====
        menubar = tk.Menu(self.root, bg="#1a2f5a", fg="#6a7a9c", activebackground="#274b87", activeforeground="#6a7a9c")
        self.root.config(menu=menubar)

        # Larger font for the menubar items
        menubar_font = ("Segoe UI", 10, "bold")

        view_menu = tk.Menu(menubar, tearoff=0, bg="white", fg="black",
                            activebackground="#274b87", activeforeground="white",
                            font=("Segoe UI", 10))

        # Add menu items
        view_menu.add_command(label="Main Page", command=self.show_main_page)
        view_menu.add_command(label="Pathway Analysis", command=self.show_pathway_page)

        # Add the "View" cascade with custom font
        menubar.add_cascade(label="  Menu  ", menu=view_menu, font=menubar_font)

        # Building UIs
        self.setup_ui()
        self.setup_pathway_ui()

    # ===== Page Switchers =====
    def show_main_page(self):
        self.pathway_frame.pack_forget()
        self.main_frame.pack(fill="both", expand=True)

    def show_pathway_page(self):
        self.main_frame.pack_forget()
        self.pathway_frame.pack(fill="both", expand=True)

    # ===== Spinner Helpers ===== (PLEASE WHY)
    def show_spinner(self):
        self.spinner_win = tk.Toplevel(self.root)
        self.spinner_win.overrideredirect(True)
        self.spinner_win.configure(bg="#0b1d3a")

        # issue here somewhere i think
        self.spinner_win.geometry("+%d+%d" % (
            self.root.winfo_rootx() + self.root.winfo_width()//2 - 75,
            self.root.winfo_rooty() + self.root.winfo_height()//2 - 25
        ))

        ttk.Label(self.spinner_win, text="Loading...", style="TLabel").pack(pady=10)
        pb = ttk.Progressbar(self.spinner_win, mode="indeterminate", length=150)
        pb.pack(pady=10, padx=20)
        pb.start(10)

        self.spinner_pb = pb
        self.spinner_win.update_idletasks()

    def hide_spinner(self):
        if hasattr(self, "spinner_win"):
            self.spinner_pb.stop()
            self.spinner_win.destroy()

    # ===== MAIN PAGE (original GUI: parent -> self.main_frame) =====
    def setup_ui(self):
        frame_top = tk.Frame(self.main_frame, bg="#0b1d3a")
        frame_top.pack(pady=10)

        tk.Button(frame_top, text="Upload Gene List File", command=self.load_gene_file,
                  bg="#1a2f5a", fg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=5)
        self.gene_label = tk.Label(frame_top, text="No file selected", bg="#0b1d3a", fg="white")
        self.gene_label.grid(row=0, column=1)

        tk.Button(frame_top, text="Upload Data File", command=self.load_data_file,
                  bg="#1a2f5a", fg="white", font=("Segoe UI", 10, "bold")).grid(row=1, column=0, padx=5)
        self.data_label = tk.Label(frame_top, text="No file selected", bg="#0b1d3a", fg="white")
        self.data_label.grid(row=1, column=1)

        frame_select = tk.Frame(self.main_frame, bg="#0b1d3a")
        frame_select.pack(pady=10)

        tk.Label(frame_select, text="Select Column A:", bg="#0b1d3a", fg="white").grid(row=0, column=0)
        self.col_a = ttk.Combobox(frame_select, width=30)
        self.col_a.grid(row=0, column=1)

        tk.Label(frame_select, text="Select Column B:", bg="#0b1d3a", fg="white").grid(row=1, column=0)
        self.col_b = ttk.Combobox(frame_select, width=30)
        self.col_b.grid(row=1, column=1)

        tk.Label(frame_select, text="Select Column C:", bg="#0b1d3a", fg="white").grid(row=2, column=0)
        self.col_c = ttk.Combobox(frame_select, width=30)
        self.col_c.grid(row=2, column=1)

        tk.Label(frame_select, text="Select Data Columns:", bg="#0b1d3a", fg="white").grid(row=3, column=0)
        self.data_cols = tk.Listbox(frame_select, selectmode=tk.MULTIPLE, width=30, height=6,
                                    bg="#1a2f5a", fg="white", selectbackground="#2e4a87",
                                    font=("Segoe UI", 10))
        self.data_cols.grid(row=3, column=1)

        tk.Button(self.main_frame, text="Process", command=self.process,
                  bg="#1a2f5a", fg="white", font=("Segoe UI", 10, "bold")).pack(pady=10)

        self.tabs = ttk.Notebook(self.main_frame)
        self.tabs.pack(expand=1, fill="both")

        self.frames = {}
        for name in ["Overlap All", "A ∩ B", "A ∩ C", "B ∩ C", "Unique to A", "Unique to B", "Unique to C"]:
            frame = tk.Frame(self.tabs, bg="#0b1d3a")
            self.frames[name] = frame
            self.tabs.add(frame, text=name)

        frame_bottom = tk.Frame(self.main_frame, bg="#0b1d3a")
        frame_bottom.pack(pady=10)

        tk.Button(frame_bottom, text="Copy Current Tab to Clipboard", command=self.copy_to_clipboard,
                  bg="#1a2f5a", fg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=5)
        tk.Button(frame_bottom, text="Download All as Excel", command=self.download_excel,
                  bg="#1a2f5a", fg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=1, padx=5)

        tk.Label(frame_bottom, text="Select Tab for Heatmap:", bg="#0b1d3a", fg="white").grid(row=1, column=0, padx=5, pady=5)
        self.source_tab_choice = ttk.Combobox(frame_bottom, values=list(self.frames.keys()), width=20)
        self.source_tab_choice.set("Overlap All")
        self.source_tab_choice.grid(row=1, column=1, padx=5, pady=5)

    def load_gene_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.show_spinner()
            self.root.after(100, lambda: self._finish_load_gene(path))

    def _finish_load_gene(self, path):
        self.gene_file = path
        self.gene_label.config(text=path.split("/")[-1])
        self.gene_df = pd.read_excel(path)
        cols = self.gene_df.columns.tolist()
        self.col_a['values'] = cols
        self.col_b['values'] = cols
        self.col_c['values'] = cols
        self.hide_spinner()

    def load_data_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.show_spinner()
            self.root.after(100, lambda: self._finish_load_data(path))

    def _finish_load_data(self, path):
        self.data_file = path
        self.data_label.config(text=path.split("/")[-1])
        self.data_df = pd.read_excel(path)
        self.data_cols.delete(0, tk.END)
        for col in self.data_df.columns:
            self.data_cols.insert(tk.END, col)
        self.hide_spinner()

    def process(self):
        try:
            selected_cols = [self.col_a.get(), self.col_b.get(), self.col_c.get()]
            selected_cols = [col for col in selected_cols if col]

            if not selected_cols:
                messagebox.showerror("Error", "Please select at least one gene list column.")
                return

            gene_sets = [set(self.gene_df[col].dropna().astype(str)) for col in selected_cols]
            selected = [self.data_cols.get(i) for i in self.data_cols.curselection()]
            if not selected:
                messagebox.showerror("Error", "Please select at least one data column.")
                return

            gene_col = self.data_df.columns[1]

            def filter_df(gene_list):
                return self.data_df[self.data_df[gene_col].astype(str).isin(gene_list)][[gene_col] + selected]

            self.results = {}

            if len(gene_sets) == 1:
                self.results["Overlap All"] = filter_df(sorted(gene_sets[0]))

            elif len(gene_sets) == 2:
                a, b = gene_sets
                self.results["A ∩ B"] = filter_df(sorted(a & b))
                self.results["Unique to A"] = filter_df(sorted(a - b))
                self.results["Unique to B"] = filter_df(sorted(b - a))

            elif len(gene_sets) == 3:
                a, b, c = gene_sets
                self.results["Overlap All"] = filter_df(sorted(a & b & c))
                self.results["A ∩ B"] = filter_df(sorted(a & b - c))
                self.results["A ∩ C"] = filter_df(sorted(a & c - b))
                self.results["B ∩ C"] = filter_df(sorted(b & c - a))
                self.results["Unique to A"] = filter_df(sorted(a - b - c))
                self.results["Unique to B"] = filter_df(sorted(b - a - c))
                self.results["Unique to C"] = filter_df(sorted(c - a - b))

            for name, df in self.results.items():
                self.display_table(name, df)

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def display_table(self, tab_name, df):
        frame = self.frames[tab_name]
        for widget in frame.winfo_children():
            widget.destroy()

        tree = ttk.Treeview(frame, columns=list(df.columns), show='headings')
        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=120)
        for _, row in df.iterrows():
            tree.insert('', 'end', values=list(row))
        tree.pack(expand=1, fill='both')

    def copy_to_clipboard(self):
        tab = self.tabs.tab(self.tabs.select(), "text")
        df = getattr(self, "results", {}).get(tab)
        if df is not None:
            pyperclip.copy(df.to_csv(sep='\t', index=False))
            messagebox.showinfo("Copied", f"{tab} data copied to clipboard.")

    def download_excel(self):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from tkinter import filedialog, messagebox

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if not save_path:
            return

 
        source_tab = self.source_tab_choice.get()
        if not hasattr(self, "results") or not self.results:
            messagebox.showerror("Error", "No results to export. Please run Process first.")
            return
        if source_tab not in self.results:
            messagebox.showerror("Error", f"'{source_tab}' not found in results.")
            return

        source_df = self.results[source_tab]
        #Broken?
        headers = list(source_df.columns) if not source_df.empty else None
        if headers is None:
            # Try to get the headers from the df
            try:
                gene_col = self.data_df.columns[1]
                selected_cols = [self.data_cols.get(i) for i in self.data_cols.curselection()]
                headers = [gene_col] + selected_cols
            except Exception:
                messagebox.showerror("Error", "Could not determine column headers for export.")
                return

        n_cols = len(headers)
        last_col_letter = self._get_column_letter(n_cols)

        # Number of genes from the chosen source tab
        n_genes = len(source_df)

        last_row = n_genes + 1

        wb = Workbook()
        wb.remove(wb.active)

        # 1) Write ALL result tabs, with headers, even if empty
        tab_order = ["Overlap All", "A ∩ B", "A ∩ C", "B ∩ C", "Unique to A", "Unique to B", "Unique to C"]
        for tab_name in tab_order:
            ws = wb.create_sheet(tab_name)
 
            for j, h in enumerate(headers, start=1):
                ws.cell(row=1, column=j, value=h)
            # Write data if present
            df = self.results.get(tab_name)
            if df is not None and not df.empty:
                for r_idx, row in enumerate(df.itertuples(index=False), start=2):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

        # Don't delete
        def q(name: str) -> str:
            return f"'{name}'" if " " in name or "∩" in name else name

        # 2) TPM to Log Transformed (headers match, values are LOG(number+1, 2) of source tab)
        ws_log = wb.create_sheet("TPM to Log Transformed")
        for j, h in enumerate(headers, start=1):
            ws_log.cell(row=1, column=j, value=h)

        for r in range(2, last_row + 1):
            
            ws_log.cell(row=r, column=1, value=f"={q(source_tab)}!A{r}")
         
            for c in range(2, n_cols + 1):
                col_letter = self._get_column_letter(c)
                ws_log.cell(
                    row=r,
                    column=c,
                    value=f"=LOG({q(source_tab)}!{col_letter}{r}+1,2)"
                )


        ws_z = wb.create_sheet("Log TPM to Z Score")
        for j, h in enumerate(headers, start=1):
            ws_z.cell(row=1, column=j, value=h)

        log_name = "TPM to Log Transformed"
        for r in range(2, last_row + 1):
     
            ws_z.cell(row=r, column=1, value=f"='{log_name}'!A{r}")
  
            row_range = f"$B{r}:${last_col_letter}{r}"
            for c in range(2, n_cols + 1):
                col_letter = self._get_column_letter(c)
                ws_z.cell(
                    row=r,
                    column=c,
                    value=(
                        f"=('{'{0}'}'!{col_letter}{r}-AVERAGE('{'{0}'}'!{row_range}))/"
                        f"STDEV('{'{0}'}'!{row_range})"
                    ).format(log_name)
                )

     
        ws_sorted = wb.create_sheet("Sorted Z Scores")
        for j, h in enumerate(headers, start=1):
            ws_sorted.cell(row=1, column=j, value=h)

        z_name = "Log TPM to Z Score"
        for r in range(2, last_row + 1):
            # Gene name
            ws_sorted.cell(row=r, column=1, value=f"='{z_name}'!A{r}")
            # Copy Z scores as-is for now
            for c in range(2, n_cols + 1):
                col_letter = self._get_column_letter(c)
                ws_sorted.cell(row=r, column=c, value=f"='{z_name}'!{col_letter}{r}")

        # Save
        wb.save(save_path)
        messagebox.showinfo("Success", f"Excel file saved to {save_path}")

    def _get_column_letter(self, idx: int) -> str:
  
        from openpyxl.utils import get_column_letter
        return get_column_letter(idx)

        # ===== PATHWAY ANALYSIS PAGE (mirrors main layout) =====
    def setup_pathway_ui(self):
        # Mapping of species to available libraries
        self.species_to_libs = {
            "Human": [
                "GO_Biological_Process_2023",
                "GO_Molecular_Function_2023",
                "GO_Cellular_Component_2023",
                "KEGG_2021_Human",
                "Reactome_2022",
                "WikiPathway_2023_Human",
                "MSigDB_Hallmark_2020"
            ],
            "Mouse": [
                "GO_Biological_Process_2023",
                "KEGG_2019_Mouse",
                "MGI_Mammalian_Phenotype_Level_4_2019",
                "WikiPathway_2019_Mouse"
            ]
        }

 
        frame_top = tk.Frame(self.pathway_frame, bg="#0b1d3a")
        frame_top.pack(pady=10)

        tk.Button(frame_top, text="Upload Gene List File", command=self.load_pathway_gene_file,
                  bg="#1a2f5a", fg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=5)
        self.pathway_gene_label = tk.Label(frame_top, text="No file selected", bg="#0b1d3a", fg="white")
        self.pathway_gene_label.grid(row=0, column=1)

      
        frame_select = tk.Frame(self.pathway_frame, bg="#0b1d3a")
        frame_select.pack(pady=10)

        tk.Label(frame_select, text="Select Gene Column:", bg="#0b1d3a", fg="white").grid(row=0, column=0, sticky="w")
        self.pathway_col = ttk.Combobox(frame_select, width=30)
        self.pathway_col.grid(row=0, column=1, padx=5)

    
        tk.Label(frame_select, text="Organism:", bg="#0b1d3a", fg="white").grid(row=1, column=0, sticky="w", pady=(6,0))
        self.pathway_org = ttk.Combobox(frame_select, width=30, values=list(self.species_to_libs.keys()))
        self.pathway_org.set("Human")  
        self.pathway_org.grid(row=1, column=1, padx=5, pady=(6,0))
        self.pathway_org.bind("<<ComboboxSelected>>", self.update_libraries)

        tk.Label(frame_select, text="Gene Set Library:", bg="#0b1d3a", fg="white").grid(row=2, column=0, sticky="w", pady=(6,0))
        self.pathway_lib = ttk.Combobox(frame_select, width=30)
        self.pathway_lib.grid(row=2, column=1, padx=5, pady=(6,0))

        # Initialize library list for default species
        self.update_libraries()

       
        tk.Button(self.pathway_frame, text="Run Pathway ORA", command=self.run_pathway_analysis,
                  bg="#1a2f5a", fg="white", font=("Segoe UI", 10, "bold")).pack(pady=10)

      
        self.ora_tabs = ttk.Notebook(self.pathway_frame)
        self.ora_tabs.pack(expand=1, fill="both")

        self.ora_frame = tk.Frame(self.ora_tabs, bg="#0b1d3a")
        self.ora_tabs.add(self.ora_frame, text="ORA Results")

        # Tree for ORA results
        cols = ["Rank", "Term", "Overlap", "P-value", "Adjusted P-value", "Odds Ratio", "Combined Score", "Genes"]
        self.ora_tree = ttk.Treeview(self.ora_frame, columns=cols, show="headings")
        for col in cols:
            self.ora_tree.heading(col, text=col)
            self.ora_tree.column(col, width=140 if col != "Genes" else 400, anchor="w")
        self.ora_tree.pack(expand=1, fill="both")

     
        frame_bottom = tk.Frame(self.pathway_frame, bg="#0b1d3a")
        frame_bottom.pack(pady=10)
        tk.Button(frame_bottom, text="Copy Results to Clipboard", command=self.copy_ora_to_clipboard,
                  bg="#1a2f5a", fg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=5)
        tk.Button(frame_bottom, text="Export Results to Excel", command=self.export_ora_to_excel,
                  bg="#1a2f5a", fg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=1, padx=5)

    def update_libraries(self, event=None):
        org = self.pathway_org.get()
        libs = self.species_to_libs.get(org, [])
        self.pathway_lib['values'] = libs
        if libs:
            self.pathway_lib.set(libs[0])

    def load_pathway_gene_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.show_spinner()
            self.root.after(100, lambda: self._finish_load_pathway_gene(path))

    def _finish_load_pathway_gene(self, path):
        try:
            self.pathway_gene_file = path
            self.pathway_gene_label.config(text=path.split("/")[-1])
            self.pathway_gene_df = pd.read_excel(path)
            cols = self.pathway_gene_df.columns.tolist()
            self.pathway_col['values'] = cols
        finally:
            self.hide_spinner()

    def run_pathway_analysis(self):
  
        if not _GSEAPY_AVAILABLE:
            messagebox.showerror(
                "gseapy not found",
                "This feature requires 'gseapy'.\nInstall with:\n\n    pip install gseapy\n\nThen relaunch the app."
            )
            return

  
        if not hasattr(self, "pathway_gene_df") or self.pathway_gene_df is None:
            messagebox.showerror("Error", "Please upload a gene list file first.")
            return
        gene_col = self.pathway_col.get()
        if not gene_col:
            messagebox.showerror("Error", "Please select a gene column.")
            return

        gene_list = (
            self.pathway_gene_df[gene_col]
            .dropna()
            .astype(str)
            .str.strip()
            .str.upper()
            .unique()
            .tolist()
        )

        if len(gene_list) == 0:
            messagebox.showerror("Error", "Selected gene column is empty after cleaning.")
            return

        library = self.pathway_lib.get() or "GO_Biological_Process_2023"
        organism = self.pathway_org.get() or "Human"

  
        try:
            self.show_spinner()
      
            enr = gp.enrichr(
                gene_list=gene_list,
                gene_sets=[library],
                organism=organism,
                outdir=None,
                no_plot=True
            )
            df = enr.results.copy() if hasattr(enr, "results") else pd.DataFrame()
        except Exception as e:
            messagebox.showerror("ORA Error", f"Failed to run ORA via gseapy:\n{e}")
            self.hide_spinner()
            return
        finally:
            self.hide_spinner()

        if df is None or df.empty:
            messagebox.showinfo("No Results", "No enriched terms were found for the selected list and library.")
     
            for item in self.ora_tree.get_children():
                self.ora_tree.delete(item)
            return

 
        keep_cols = ["Rank", "Term", "Overlap", "P-value", "Adjusted P-value", "Odds Ratio", "Combined Score", "Genes"]
        for item in self.ora_tree.get_children():
            self.ora_tree.delete(item)

 
        if "Rank" not in df.columns:
            df = df.reset_index(drop=True)
            df.insert(0, "Rank", df.index + 1)

        # Ensure all columns exist
        for c in keep_cols:
            if c not in df.columns:
                df[c] = ""

        # Sort by Adjusted P-value when present
        try:
            df = df.sort_values(by="Adjusted P-value", ascending=True)
        except Exception:
            pass

        # Save for export/copy
        self.ora_results_df = df[keep_cols].copy()

        # Insert into tree
        for _, row in self.ora_results_df.iterrows():
            vals = [row.get(c, "") for c in keep_cols]
            # Coerce to str for treeview
            vals = [str(v) for v in vals]
            self.ora_tree.insert("", "end", values=vals)

    def copy_ora_to_clipboard(self):
        df = getattr(self, "ora_results_df", None)
        if df is None or df.empty:
            messagebox.showerror("Error", "No ORA results to copy. Please run ORA first.")
            return
        pyperclip.copy(df.to_csv(sep="\t", index=False))
        messagebox.showinfo("Copied", "ORA results copied to clipboard.")

    def export_ora_to_excel(self):
        df = getattr(self, "ora_results_df", None)
        if df is None or df.empty:
            messagebox.showerror("Error", "No ORA results to export. Please run ORA first.")
            return
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if not save_path:
            return
        try:
            df.to_excel(save_path, index=False)
            messagebox.showinfo("Success", f"ORA results saved to {save_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save ORA results:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = GeneComparerApp(root)
    root.mainloop()
